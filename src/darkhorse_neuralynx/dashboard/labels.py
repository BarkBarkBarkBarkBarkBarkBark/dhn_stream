"""Channel label loading for the optional dashboard."""

from __future__ import annotations

import csv
import re
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET


@dataclass(frozen=True)
class ChannelLabel:
    channel: int
    name: str
    description: str = ""
    electrode_type: str = "unknown"
    source: str = "fallback"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _normalise_key(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def load_channel_config_csv(path: str | Path) -> dict[int, ChannelLabel]:
    """Load DHN_Acq channel names/descriptions from a CSV or TSV channel settings file."""
    p = Path(path)
    delimiter = "\t" if p.suffix.lower() == ".tsv" else ","
    labels: dict[int, ChannelLabel] = {}
    with p.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            return labels
        field_map = {_normalise_key(name): name for name in reader.fieldnames}
        number_key = field_map.get("channel_number")
        name_key = field_map.get("channel_name")
        description_key = field_map.get("channel_description")
        if number_key is None:
            raise ValueError(f"No Channel Number column found in {p}")
        for row in reader:
            raw_number = (row.get(number_key) or "").strip()
            if not raw_number:
                continue
            try:
                channel = int(raw_number)
            except ValueError:
                continue
            name = (row.get(name_key) or f"ch{channel:04d}").strip() if name_key else f"ch{channel:04d}"
            description = (row.get(description_key) or "").strip() if description_key else ""
            labels[channel] = ChannelLabel(
                channel=channel,
                name=name or f"ch{channel:04d}",
                description=description,
                source=str(p),
            )
    return labels


def load_connection_map(path: str | Path) -> dict[int, ChannelLabel]:
    """Load channel labels from CSV/TSV/XLSX connection maps using best-effort columns."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _load_connection_map_rows(p, _read_csv_rows(p))
    if suffix == ".xlsx":
        return _load_connection_map_xlsx(p)
    raise ValueError(f"Unsupported connection map format: {p.suffix}")


def merge_labels(
    n_channels: int,
    *sources: dict[int, ChannelLabel],
) -> list[ChannelLabel]:
    """Merge label sources by precedence: later sources override earlier sources."""
    merged: dict[int, ChannelLabel] = {}
    for channel in range(1, n_channels + 1):
        merged[channel] = ChannelLabel(channel=channel, name=f"ch{channel:04d}")
    for source in sources:
        for channel, label in source.items():
            if 1 <= channel <= n_channels:
                merged[channel] = label
    return [merged[channel] for channel in range(1, n_channels + 1)]


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def _load_connection_map_xlsx(path: Path) -> dict[int, ChannelLabel]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _load_connection_map_xlsx_stdlib(path)

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        all_rows: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = _find_header_index(rows)
            if header_index is None:
                continue
            headers, data_start = _headers_and_data_start(rows, header_index)
            for row in rows[data_start:]:
                record = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)}
                record["_sheet"] = sheet.title
                all_rows.append(record)
        return _load_connection_map_rows(path, all_rows)
    finally:
        workbook.close()


def _load_connection_map_xlsx_stdlib(path: Path) -> dict[int, ChannelLabel]:
    rows_by_sheet = _read_xlsx_rows_stdlib(path)
    all_rows: list[dict[str, Any]] = []
    for sheet_name, rows in rows_by_sheet.items():
        header_index = _find_header_index(rows)
        if header_index is None:
            continue
        headers, data_start = _headers_and_data_start(rows, header_index)
        for row in rows[data_start:]:
            record = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)}
            record["_sheet"] = sheet_name
            all_rows.append(record)
    return _load_connection_map_rows(path, all_rows)


def _read_xlsx_rows_stdlib(path: Path) -> dict[str, list[tuple[Any, ...]]]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with zipfile.ZipFile(path) as archive:
        shared_strings = _read_shared_strings(archive, ns)
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in relationships.findall("r:Relationship", rel_ns)
        }
        sheets: dict[str, list[tuple[Any, ...]]] = {}
        for sheet in workbook.findall("x:sheets/x:sheet", ns):
            name = sheet.attrib.get("name", "Sheet")
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if rel_id is None or rel_id not in rel_targets:
                continue
            target = rel_targets[rel_id]
            sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            tree = ET.fromstring(archive.read(sheet_path))
            sheets[name] = _sheet_rows(tree, shared_strings, ns)
        return sheets


def _read_shared_strings(archive: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    tree = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in tree.findall("x:si", ns):
        text_parts = [node.text or "" for node in item.findall(".//x:t", ns)]
        strings.append("".join(text_parts))
    return strings


def _sheet_rows(tree: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for row in tree.findall(".//x:sheetData/x:row", ns):
        cells: dict[int, Any] = {}
        for cell in row.findall("x:c", ns):
            ref = cell.attrib.get("r", "A1")
            idx = _column_index(ref)
            cells[idx] = _cell_value(cell, shared_strings, ns)
        if cells:
            max_idx = max(cells)
            rows.append(tuple(cells.get(idx) for idx in range(max_idx + 1)))
    return rows


def _cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        text = cell.find(".//x:t", ns)
        return text.text if text is not None else ""
    value = cell.find("x:v", ns)
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        idx = int(value.text)
        return shared_strings[idx] if idx < len(shared_strings) else ""
    try:
        number = float(value.text)
    except ValueError:
        return value.text
    return int(number) if number.is_integer() else number


def _column_index(cell_ref: str) -> int:
    letters = re.match(r"[A-Z]+", cell_ref.upper())
    if letters is None:
        return 0
    idx = 0
    for char in letters.group(0):
        idx = idx * 26 + (ord(char) - ord("A") + 1)
    return idx - 1


def _find_header_index(rows: list[tuple[Any, ...]]) -> int | None:
    for idx, row in enumerate(rows[:20]):
        keys = {_normalise_key(str(value)) for value in row if value is not None}
        if any("channel" in key or key in {"ch", "chan"} for key in keys):
            return idx
    return None


def _headers_and_data_start(rows: list[tuple[Any, ...]], header_index: int) -> tuple[list[str], int]:
    base = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    next_row = rows[header_index + 1] if header_index + 1 < len(rows) else ()
    sub = [str(value).strip() if value is not None else "" for value in next_row]
    has_subheaders = any(value.lower() in {"start", "end"} for value in sub)
    headers: list[str] = []
    current_parent = ""
    width = max(len(base), len(sub))
    for idx in range(width):
        parent = base[idx] if idx < len(base) else ""
        child = sub[idx] if idx < len(sub) else ""
        if parent:
            current_parent = parent
        if has_subheaders and child:
            headers.append(f"{current_parent} {child}".strip())
        else:
            headers.append(parent or child or f"column_{idx + 1}")
    return headers, header_index + (2 if has_subheaders else 1)


def _load_connection_map_rows(path: Path, rows: list[dict[str, Any]]) -> dict[int, ChannelLabel]:
    labels: dict[int, ChannelLabel] = {}
    for row in rows:
        normalised = {_normalise_key(str(key)): value for key, value in row.items() if key is not None}
        channel = _first_int(normalised, ("channel", "channel_number", "ch", "chan", "dhn_channel"))
        range_start = _first_int(
            normalised,
            ("channel_range_(csc#)_start", "channel_range_csc_start", "csc_start", "start"),
        )
        range_end = _first_int(
            normalised,
            ("channel_range_(csc#)_end", "channel_range_csc_end", "csc_end", "end"),
        )
        if channel is None:
            if range_start is not None and range_end is not None:
                _add_connection_range(labels, path, normalised, range_start, range_end)
            continue
        name = _first_text(normalised, ("label", "name", "channel_name", "electrode", "contact"))
        if not name:
            name = _first_text(normalised, ("electrode_label",))
        description = _first_text(normalised, ("description", "notes", "region", "location", "target"))
        if not description:
            description = _first_text(normalised, ("electrode_description",))
        electrode_type = _infer_electrode_type(" ".join(str(value) for value in normalised.values()))
        labels[channel] = ChannelLabel(
            channel=channel,
            name=name or f"ch{channel:04d}",
            description=description,
            electrode_type=electrode_type,
            source=str(path),
        )
    return labels


def _add_connection_range(
    labels: dict[int, ChannelLabel],
    path: Path,
    row: dict[str, Any],
    start: int,
    end: int,
) -> None:
    if end < start:
        start, end = end, start
    base_name = _first_text(row, ("electrode_label", "label", "name", "electrode"))
    description = _first_text(row, ("electrode_description", "description", "region", "location", "target"))
    electrode_type = _infer_electrode_type(" ".join(str(value) for value in row.values()))
    for offset, channel in enumerate(range(start, end + 1), start=1):
        labels[channel] = ChannelLabel(
            channel=channel,
            name=f"{base_name}{offset}" if base_name else f"ch{channel:04d}",
            description=description,
            electrode_type=electrode_type,
            source=str(path),
        )


def _first_int(row: dict[str, Any], candidates: tuple[str, ...]) -> int | None:
    for key in candidates:
        value = row.get(key)
        if value is None or value == "":
            continue
        try:
            return int(float(str(value).strip()))
        except ValueError:
            continue
    return None


def _first_text(row: dict[str, Any], candidates: tuple[str, ...]) -> str:
    for key in candidates:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _infer_electrode_type(text: str) -> str:
    lowered = text.lower()
    if "micro" in lowered:
        return "micro"
    if "macro" in lowered:
        return "macro"
    return "unknown"
